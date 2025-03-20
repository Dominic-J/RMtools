# -*- coding: utf-8 -*-
"""This module provide some useful tools for universal mission.
Created on Thu May 28 15:06:51 2015 @author: 左词"""
import numpy as np
import pandas as pd
from sqlalchemy import create_engine
from . import config
from datetime import datetime,timedelta ,date
#from psycopg2 import connect
import shelve as sv
import inspect
import sys
import re
import string
import os
from openpyxl.styles import Alignment ,Font ,Border ,Side ,PatternFill ,GradientFill
from openpyxl import load_workbook


def obj_info(obj):
    """查看对象的方法、属性列表，并归到不同的类别中去

    参数:
    ---------
    obj : 任意 python 对象，最常见的输入值种类有模块、类、对象

    返回值:
    ---------
    obj_info : ObjInfo 对象，自定义对象，承载有 obj 的属性、函数、方法、模块等不同的属性类别清单
        obj_info.modules 属性保存 obj 的模块清单
        obj_info.classes 属性保存 obj 的类、对象清单
        obj_info.functions 属性保存 obj 的函数、方法清单
        obj_info.attributes 属性保存 obj 的狭义的属性清单"""

    class ObjInfo:
        """自定义对象，承载一个 python 对象的属性、函数、方法、模块等不同的属性类别清单
        obj_info.modules 属性保存 obj 的模块清单
        obj_info.classes 属性保存 obj 的类、对象清单
        obj_info.functions 属性保存 obj 的函数、方法清单
        obj_info.attributes 属性保存 obj 的狭义的属性清单"""
        def __init__(self):
            self.modules = []
            self.classes = []
            self.functions = []
            self.attributes = []

        def __iter__(self):
            attr_list = self.modules + self.classes + self.functions + self.attributes
            return iter(attr_list)   # __iter__方法已经返回了可迭代对象，因此不需要再实现__next__方法

        def __repr__(self):
            strp = "ObjInfo object of :\n"
            if self.modules:
                strp += "    模块：{}\n\n".format(str(self.modules))
            if self.classes:
                strp += "    类/对象：{}\n\n".format(str(self.classes))
            if self.functions:
                strp += "    函数/方法：{}\n\n".format(str(self.functions))
            if self.attributes:
                strp += "    属性：{}\n\n".format(str(self.attributes))
            return strp

        def __str__(self):
            return self.__repr__()

        def iterms_count(self):
            """计算各个类别（模块、类/对象、函数/方法、属性）中，分别有多少个元素。返回值是字典"""
            cnt = {'total': 0}
            cnt['modules'] = len(self.modules)
            cnt['total'] += len(self.modules)

            cnt['classes'] = len(self.classes)
            cnt['total'] += len(self.classes)

            cnt['functions'] = len(self.functions)
            cnt['total'] += len(self.functions)

            cnt['attributes'] = len(self.attributes)
            cnt['total'] += len(self.attributes)
            return cnt

    from inspect import isclass, ismethod, isfunction, ismodule

    attr = [i for i in dir(obj) if not i.startswith('_')]
    obj_info = ObjInfo()
    for name in attr:
        obj_attr = getattr(obj, name)
        if ismodule(obj_attr):
            obj_info.modules.append(name)
            continue
        if isclass(obj_attr):
            obj_info.classes.append(name)
            continue
        if ismethod(obj_attr) or isfunction(obj_attr):
            obj_info.functions.append(name)
            continue
        obj_info.attributes.append(name)
    return obj_info


def class_tree(cls, indent=3):
    """打印出类的继承树出来，缩进量越多的类，是层级越高（越早）的超类

    参数:
    ----------
    cls: 任意的类"""
    print('.' * indent + cls.__name__)
    for supercls in cls.__bases__:
        class_tree(supercls, indent + 3)


def instance_tree(inst):
    """打印出对象的继承树出来，缩进量越多的类，是层级越高（越早）的超类

    参数:
    ----------
    inst: 任意实例化的对象"""
    print('Tree of %s' % inst)
    class_tree(inst.__class__, 3)


def ordered_diff(col1, col2):
    """同 set(col1) - set(col2)，计算 col1 与 col2的元素差集，但会保持原列表中的元素次序

    参数:
    --------
    col1 : 任意支持迭代、同时有序的元素集合类对象，比如 list, Series
    col2 : 任意支持迭代、同时有序的元素集合类对象，比如 list, Series. col2 不必与 col1 类型相同

    返回值:
    --------
    diff_list : list, col1 与 col2 的元素差集"""
    return [i for i in col1 if i not in col2]


def ordered_interact(col1, col2):
    """同 set(col1) & set(col2)，计算 col1 与 col2的元素交集，但会保持原列表中的元素次序

    参数:
    --------
    col1 : 任意支持迭代、同时有序的元素集合类对象，比如 list, Series
    col2 : 任意支持迭代、同时有序的元素集合类对象，比如 list, Series. col2 不必与 col1 类型相同

    返回值:
    --------
    diff_list : list, col1 与 col2 的元素差集"""
    return [i for i in col1 if i in col2]


def check_diff(sr1, sr2, detail=False):
    """检查两列的差集。拼表前检查数据质量时，经常需要对拼接键做这种检查。  \n
    此函数也可用来检查 df1, df2 的 columns 差集、df1.index 与 df2.index 差集。  \n
    参数:
    ---------
    sr1: iterable, 列1  \n
    sr2: iterable, 列2， sr1 和 sr2 是检查差集的目标对象  \n
    detail: bool, 是否要返回差集的明细 diff1, diff2 。默认只打印出差集中的元素个数，不返回差集明细  \n
    """
    diff1 = set(sr1) - set(sr2)
    diff2 = set(sr2) - set(sr1)
    print('diff1: {}'.format(len(diff1)))
    print('diff2: {}'.format(len(diff2)))
    if detail:
        return diff1, diff2
    
    
def re_search(patten, str_set):
    """用正则表达式，在由字符串组成的集合中搜索，返回所有匹配的搜索结果\n

    参数:
    ----------
    patten: str, 正则表达式描述的模式 \n
    str_set: iterable of str, 搜索的目标集合 \n

    返回值:
    ----------
    match: list, 包含所有匹配模式的item
    """
    from re import compile
    patten = compile(patten)
    match = []
    for i in str_set:
        tmp = patten.search(i)
        if tmp is not None:
            match.append(tmp.string)
    return match


def doc_search(patten, obj):
    """从任意对象的所有属性、方法、类等对象的文档说明中，查找是否有符合
    给定模式的文档说明。若有，则返回该属性名字

    参数:
    ------------
    patten: str, 用正则表达式描述的模式，即你希望查找的模式
    obj: 任意 python 对象

    返回:
    ------------
    attribute_list: 符合 patten 描述的所有 obj 的属性清单
    """

    b = obj_info(obj)
    results = []
    for name in b:
        obj_in_pc = getattr(obj, name)
        doc = getattr(obj_in_pc, '__doc__')
        if doc:  # 如果有文档
            result = re_search(patten, [doc])
            if result:
                results.append(name)
    return results


def reverse_dict(d):
    """把字典 d 中的 key:value 对反转为 value:key , 返回另一字典. value 必须是不可变的，否则会报错  \n"""
    return {value: key for key, value in d.items()}


def foreach(func, iterator):
    """依次对iterator中的所有元素应用func函数。没有返回值\n

    参数:
    ----------
    func: 任意函数，以iterator中的元素为输入参数，无返回值 \n
    iterator: 任意可迭代对象"""
    for item in iterator:
        func(item)


def print_each(iterator, format_str=None):
    """依次打印集合元素中的元素 \n

    参数:
    ----------
    iterator: 任意可迭代对象
    format_str: str or None, 可选
        用于格式化打印每个元素的格式化字符串。

    示例:
    ----------
    > format_str = '{:02d}'
    > print_each(range(1,5), format_str)   # 打印结果如下
    01
    02
    03
    04"""
    for item in iterator:
        if format_str is not None:
            print(format_str.format(item))
        else:
            print(item)


def dical_prod(set1, set2, return_type='string'):
    """把 set1 和 set2 中的元素做笛卡尔积组合。如果 set1 有 m 个元素，set2 有 n 个元素，则组合结果有 m*n个 元素 \n

    参数:
    ----------
    set1, set2 : 任意可迭代的对象，需要组合的两个对象。  \n
    return_type: str, 'string' 表示把 2 个元素合并成一个字符串，以'list' 表示把 2 个元素放在一个 list 中  \n """
    dical = [[i, j] for i in set1 for j in set2]
    if return_type == 'string':
        dical = [item[0] + '_' + item[1] for item in dical]
    return dical


def flatten(x):
    """展平嵌套列表结构"""
    out = []
    for i in x:
        if isinstance(i, (list, tuple)):
            out.extend(flatten(i))
        else:
            out.append(i)
    return out


def subset(super_set):
    """返回 super_set 的所有子集"""
    super_set = set(super_set)
    assert len(super_set) >= 2, 'supSet has only one element, so it has no subset'
    sub1 = set()
    sub_list = set()
    for i in super_set:
        sub1.add(frozenset(super_set - {i}))
        sub_list.update(sub1)
    if len(super_set) == 2:
        return [set(i) for i in sub_list]
    else:
        for i in sub1:
            tmp = subset(i)
            tmp = [frozenset(i) for i in tmp]
            sub_list.update(tmp)
        return [set(i) for i in sub_list]


#%% 日期、时间相关
def date2str(date_obj, sep='-'):
    """把datetime.date对象转换成日期字符串, 转换的格式由 sep 参数决定  \n

    参数:
    ----------
    date_obj: datetime.date or datetime.datetime 对象.  \n
    sep: 分隔符，指定转换成什么格式的日期字符串:   \n
        '-' 表示转换成 '2017-06-01' 格式的日期  \n
        '/' 表示转换成 '2017/06/01' 格式的日期  \n
        ''  表示转换成 '20170601' 格式的日期

    返回值:
    ----------
    date_str: str, 日期字符串"""

    assert sep in ('-', '/', ''), "仅支持sep 在('-', '/', '')中取值"
    str_f = '%Y{0}%m{0}%d'.format(sep)
    date_str = date_obj.strftime(str_f) if pd.notnull(date_obj) else 'NaT'
    return date_str


def str2date(date_str):
    """把日期字符串转换成 datetime.date对象。  \n

     参数:
     ------------
     date_str: str, 表示日期的字符串。  \n
        只要date_str的前 8/10 位表示日期就行，多余的字符串不影响解析； \n
        日期格式只能是下列之一： \n
            '2017-06-01' 格式的日期,   \n
            '2017/06/01' 格式的日期；  \n
            '20170601' 格式的日期。   \n

    返回值:
    -----------
    date_obj: datetime.date 对象。"""
    if len(str(date_str)) >= 10 and date_str is not None:
        tmp = date(int(str(date_str)[:4]) , int(str(date_str)[5:7]) , int(str(date_str)[8:10]))        
    elif len(str(date_str)) == 8 and date_str is not None:
        tmp = date(int(str(date_str)[:4]) , int(str(date_str)[4:6]) , int(str(date_str)[6:8]))
    else:
        tmp = pd.NaT
    return tmp


def str2datetime(str_datetime):
    """把字符串形式的datetime, 转换成 datetime 类型。\n

    参数:
    -----------
    str_datetime: str, 2018-03-12 12:28:32 格式的时间
    
    返回值:
    -----------
    dtime: datetime, 转换后的 datetime 对象"""
    return datetime.strptime(str_datetime, '%Y-%m-%d %H:%M:%S')


def date_period(time_sr, freq='week'):
    """根据日期列，生成对应的周/月/季/年，用来按周/月/年汇总其他数据。   \n

    参数:
    -----------
    time_sr: series,日期字符串('2017-06-01' 或 '2017/06/01' 或 '20170601' 格式，多余字符串不影响） \n
        或者 datetime.date/datetime.datetime 对象   \n
    freq: str, 期望转换成的周期，可选取值有 'week', 'month', 'quarter' or 'year'  \n  
    
    返回值:
    -----------
    period_sr: series_of_str, 返回对应的周期序列。"""

    from datetime import timedelta
    if freq.upper() == 'WEEK':
        # 字符串格式的日期，先转换成 datetime.date
        if isinstance(time_sr.iloc[0], str):
            date_str = time_sr.iloc[0]
            sep = date_str[4]
            if sep not in ('-', '/'):
                sep = ''
            time_sr = time_sr.apply(str2date, sep=sep)
            
        fun = lambda x: (x - timedelta(days=x.weekday())).strftime('%Y-%m-%d') if pd.notnull(x) else 'NaT'
        return time_sr.apply(fun)

    elif freq.upper() == 'MONTH':
        # 字符串格式的日期
        if isinstance(time_sr.iloc[0], str):
            date_str = time_sr.iloc[0]
            sep = date_str[4]
            if sep == '-':
                return time_sr.apply(lambda x: x[:7] if pd.notnull(x) else 'NaT')
            elif sep == '/':
                return time_sr.apply(lambda x: x[:7].replace('/', '-') if pd.notnull(x) else 'NaT')
            else:  # '20170601' 格式的日期串
                return time_sr.apply(lambda x: x[:4] + '-' + x[4:6] if pd.notnull(x) else 'NaT')
        else:  # date 或 datetime 对象
            return time_sr.apply(lambda x: x.strftime('%Y-%m') if pd.notnull(x) else 'NaT')
    
    elif freq.upper() == 'QUARTER':
        # 字符串格式的日期
        if isinstance(time_sr.iloc[0], str):
            date_str = time_sr.iloc[0]
            sep = date_str[4]
            month = lambda x: int(x[5:7]) if sep in ('-', '/') else lambda x: int(x[4:6])
            quarter = lambda x: (month(x) - 1) // 3 + 1
            return time_sr.apply(lambda x: '{year}Q{q}'.format(year=x[:4], q=quarter(x))
                                 if pd.notnull(x) else 'NaT')
        else:  # data 或 datetime对象
            quarter = lambda x: (x.month - 1) // 3 + 1
            return time_sr.apply(lambda x: '{year}Q{q}'.format(year=x.year, q=quarter(x))
                                 if pd.notnull(x) else 'NaT')
        
    elif freq.upper() == 'YEAR':
        if isinstance(time_sr.iloc[0], str):
            return time_sr.apply(lambda x: x[:4] if pd.notnull(x) else 'NaT')
        else:
            return time_sr.apply(lambda x: x.strftime('%Y') if pd.notnull(x) else 'NaT')

    else:
        raise Exception("Unkown freq {}: it should be in ('week','month','quarter','year')".format(freq))


def month_diff(date1_sr, date2_sr, ignore_days=True):
    """两个日期的月份差
    参数:
    ------------
    date1_sr: series of str or date,
        若是字符串日期，格式需为：yyyy-mm-dd 或 yyyy/mm/dd 或 yyyymmdd
    date2: 同 date1_sr, date2 不必与 date1_sr 格式相同，但需符合规范
    ignore_days: bool, 是否需要考虑日的精度，即满整月与否。
    
    返回:
    ------------
    months: series, 月份差"""

    def base(row):
        """以表格的一行为输入，计算月份差"""
        from datetime import date
        if isinstance(row.iloc[0],str):
            if len(row.iloc[0]) == 10:
                date1 = str2date(row.iloc[0])
            else:
                date1 = str2date(row.iloc[0]+'-01')
        elif pd.isnull(row.iloc[0]) or row.iloc[0] == None:
            date1 = pd.NaT
        else:
            date1 = row.iloc[0]
            
        if isinstance(row.iloc[1],str):
            if len(row.iloc[1]) == 10:
                date2 = str2date(row.iloc[1])
            else:
                date2 = str2date(row.iloc[1]+'-01')
        elif pd.isnull(row.iloc[1]) or row.iloc[1] == None:
            date2 = pd.NaT
        else:
            date2 = row.iloc[1]
    
        diff = (date1.year - date2.year) * 12 + date1.month - date2.month
        if not ignore_days:
            if date1.days < date2.days:
                diff -= 1

        return diff

    date_df = pd.concat([date1_sr, date2_sr], axis=1)
    return date_df.apply(base, axis=1)


#%% 读、写相关
def to_pickle(file_name ,obj):
    """将任意 python 对象写成 pickle 文件。 函数无返回

    参数:
    ----------
    obj: 任意待写到磁盘的 python 对象
    path: str, 写入的目标文件的绝对路径，不需要带扩展名
    """
    if file_name.find('/') > -1 or file_name.find('\\') > -1:
        file_path = file_name + '.pkl'
    else:
        if sys.platform == 'linux':
            file_path = r'//data/result_data/pickle/' + file_name + '.pkl'
        elif sys.platform == 'win32':
            file_path = r'D:/Result_data/pickle/' + file_name + '.pkl'
        else:
            pass
    from pickle import dump
    with open(file_path, 'wb') as f:
        dump(obj, f)


def read_pickle(file_name):
    """读取任意 pickle 对象

    参数:
    ----------
    path: str, 写入的目标文件的绝对路径，不需要带扩展名

    返回:
    ----------
    obj: 任意 python 对象"""
    from pickle import load      
    if file_name.find('/') > -1 or file_name.find('\\') > -1:
        file_path = file_name
    else:
        if sys.platform == 'linux':
            file_path = r'//data/result_data/pickle/' + file_name + '.pkl'
        elif sys.platform == 'win32':
            file_path = r'D:/Result_data/pickle/' + file_name + '.pkl'
        else:
            pass
    with open(file_path, 'rb') as f:
        obj = load(f)
    return obj


# shelve 增强
def write(self, var_name):
    """把名为 var_name 的变量写入shelve文件中。当要保存多个变量时，使用此方法更方便\n

    参数:
    -----------
    var_name: str or list_of_str, 需要保存数据的变量的名字\n

    示例:
    -----------
    self.write(['a','b']) 等同于下列代码： \n
    self['a'] = a \n
    self['b'] = b """
    if isinstance(var_name, str):   # 如果是单个变量
        var_name = [var_name]
    for i in var_name:
        exec('self["{0}"] = {0}'.format(i))
sv.DbfilenameShelf.write = write


def dec(fun):
    def keys(self):
        """返回由键组成的list"""
        k = fun(self)  # fun 是旧的sv.DbfilenameShelf.keys函数
        return list(k)  # k 是个KeyView，无法直接查看具体的items
    return keys
sv.DbfilenameShelf.keys = dec(sv.DbfilenameShelf.keys)
del dec


def varname2str(var):
    """
    将变量名称转化为字符串
    """
    for fi in reversed(inspect.stack()):
        names = [var_name for var_name, var_val in fi.frame.f_locals.items() if var_val is var]
        if len(names) > 0:
            return names[0]


def to_excels(file_name , *df_list ,sheet_name=None ,color_dict=None ,color_num = 3,index=False ,drop_sheet=True ):
    """把多个df的数据，写入指定的同一个excel文件中去。本函数无返回值。

    参数：
    ---------
    file_name: str，指定的 excel 文件名，不需要带扩展名，默认存储在  //data/result_data/excel 、D:/Result_data/excel/ ，若无则带完整路径名
    df_list: dataframe, 任意个数的 dataframe 表
    sheet_name :每个dataframe写入excel对应的名字 ，可以是str or list
    index：是否删除索引
    drop_sheet：是否覆盖已经存在 sheet_name的表 ，默认删除
    color_dict: 是否根据这一列做图表的视觉效果，目前只支持针对某一列进行颜色填充
    color_num: 做视觉效果的颜色种类 ，默认三种颜色，最多五种颜色，目前颜色不可选
    """
    def getColmunName(index):
        """
        得到excel的列所在的名称如  A ,B ,C ,D等
        """
        ret = ''
        ci = index - 1
        col_row = ci // 26
        if col_row > 0:
            ret += getColmunName(col_row)
        ret += string.ascii_uppercase[ci % 26]
        return ret
    
    if file_name.find('/') > -1 or file_name.find('\\') > -1:
        file_path = file_name + '.xlsx'
    else:   
        if sys.platform == 'linux':
            file_path = r'//data/result_data/excel/' + file_name + '.xlsx'
        elif sys.platform == 'win32':
            file_path = r'D:/Result_data/excel/' + file_name + '.xlsx'
        else:
            pass
    if os.path.exists(file_path):
        EW = pd.ExcelWriter(file_path ,mode='a' ,engine='openpyxl')
    else:
        EW = pd.ExcelWriter(file_path)
    with EW as writer:    # doctest: +SKIP
        for i, df in enumerate(df_list):
            if isinstance(df,pd.DataFrame):
                if isinstance(color_dict ,str):
                    dfinl = df.sort_values(color_dict).copy()
                    dfi = dfinl.drop_duplicates(color_dict)[color_dict].to_frame().reset_index(drop=True).reset_index()
                    dfi['rak'] = dfi['index'].apply(lambda x: (x+1)%color_num)
                    dfinl = dfinl.merge(dfi[[color_dict,'rak']] ,on = color_dict)
                    dfinl.to_excel(writer, sheet_name='{}'.format("Sheet" + str(i)) ,index=index)
                df.to_excel(writer, sheet_name='{}'.format("Sheet" + str(i)) ,index=index)
    for f, df in enumerate(df_list):
        col_n = [len(str(i)) - len([j for j in str(i) if '\u4e00' <= j <= '\u9fff']) + 
                 len([j for j in str(i) if '\u4e00' <= j <= '\u9fff'])*2 for i in df.col_list()]
        col_m = [i for i in df.col_list()]
        dict_col = {}
        dfn = pd.DataFrame()
        for i in range(len(col_n)):
            dict_col[col_m[i] + '_len'] = col_n[i]
            dfn[col_m[i] + '_len'] = df[col_m[i]].apply(lambda x: len([i for i in str(x)]) - 
                            len([i for i in str(x) if '\u4e00' <= i <= '\u9fff']) + len([i for i in str(x) if '\u4e00' <= i <= '\u9fff'])*2)
        dfi_colslen = pd.DataFrame(dict_col ,index=[0])

        dfi_colslen = pd.concat([dfi_colslen ,dfn])
        cols_max = []
        for j in col_m:
            vmn = round((dfi_colslen[j + '_len'].max()+1),1)
            if vmn < 4: vmn+=1
            cols_max.append(vmn)

        side = Side(style='thin' ,color = '000000')
        border = Border(top=side ,bottom=side ,left=side ,right=side)
        font = Font(name='微软雅黑' ,size = 10.5 ,bold = True ,color = 'FFFFFF')
        font1 = Font(name='微软雅黑' ,size = 9 ,color = '000000')
        gradfill = GradientFill(degree=60 ,stop=("1E90FF","1E90FF"))

        wb = load_workbook(file_path)
        ws = wb['Sheet'+str(f)]
        for i in range(len(cols_max)):
            m = getColmunName(i+1)
            if cols_max[i] >= 50:
                tp = 35
            else:
                tp = cols_max[i]
            ws.column_dimensions[m].width = tp #设置自适应行高
        for row in ws.iter_rows(min_row=ws.min_row+1 ,max_row= ws.max_row ,min_col= ws.min_column ,max_col=ws.max_column):
            for cell in row:
                cell.font = font1
                cell.border = border
                cell.alignment = Alignment(wrap_text=True ,horizontal='left' ,vertical='center')
        for row in ws.iter_rows(min_row=1 ,max_row= 1 ,min_col= ws.min_column ,max_col=ws.max_column):
             for cell in row:
                cell.font = font
                cell.border = border
                cell.alignment = Alignment(wrap_text=True ,horizontal='center' ,vertical='center')   
                cell.fill = gradfill  
        
        wb.save(file_path)         
    if isinstance(color_dict ,str):
        """
        是否做视觉冲击的效果
        """
        wb = load_workbook(file_path)
        rk = list(set(dfinl['rak'].to_list()))
        for i in rk:
            cols_don = dfinl.loc[dfinl['rak']==i][color_dict].drop_duplicates().to_list()
            for j in cols_don:
                mf = dfinl.loc[dfinl[color_dict] == j].index
                for row in ws.iter_rows(min_row=mf.min()+2 ,max_row= mf.max()+2 ,min_col= ws.min_column ,max_col=ws.max_column):
                    for cell in row:
                        if i == 0:
                            cell.fill = GradientFill(degree=60 ,stop=("FFFACD","FFFACD"))
                        elif i == 1:
                            cell.fill = GradientFill(degree=60 ,stop=("FFB6C1","FFB6C1"))
                        elif i == 2:
                            cell.fill = GradientFill(degree=60 ,stop=("99FF33","99FF33"))
                        elif i == 3:
                            cell.fill = GradientFill(degree=60 ,stop=("99FFFF","99FFFF"))
                        elif i == 4:
                            cell.fill = GradientFill(degree=60 ,stop=("FFB3FF","FFB3FF"))
                        else:
                            pass
        
        wb.save(file_path)
        
    if isinstance(sheet_name,(str,list)):
        """
        更改表名称
        """

        if isinstance(sheet_name ,str):sheet_name = [sheet_name]
        sheet_names = wb.sheetnames #获取工作博所有的工作表名
        cols_int = [int(i[5:])  for i in sheet_names  if i.find("Sheet") > -1]
        cols_drop = [i for i in sheet_name if i in sheet_names]
        if drop_sheet:
            for i in cols_drop:
                if isinstance(i,str):
                    wb.remove(wb[i])
        for j in range(min(len(cols_int),len(sheet_name))):
            ws = wb['Sheet'+str(j)]
            ws.title = sheet_name[j]
        wb.save(file_path)                  


def to_csv(file_name ,df):
    """做CSV文件的存储。

    参数：
    ---------
    file_name: str，指定的 excel 文件名，不需要带扩展名，默认存储在  //data/result_data/excel 、D:/Result_data/excel/ ，若无则带完整路径名
    df_list: dataframe
    """
    if file_name.find('/') == -1 or file_name.find('\\') == -1:
        if sys.platform == 'linux':
            file_path = r'//data/result_data/excel/'
        elif sys.platform == 'win32':
            file_path = r'E:/yangfm/A0excel/'
        else:
            pass
        path = file_path + file_name+'.csv'
    else:
        path = file_name+'.csv'
    df.to_csv(path,index=False ,sep=',',encoding='utf-8')


class DBinfo:
    __doc__ = """把数据库的元信息读出用于搜索了解。主要信息有：数据库的表、视图，每个表的字段详情。\n
    
    参数:
    -----
    db_name: 
        str, 数据库连接名称。查看已配置了哪些数据库连接：
        MysqlTools.print_connects()
    
        数据库连接可在 config 模块中配置 """

    def __init__(self, db_name):
        url = config.MYSQL_DICT[db_name]
        con = create_engine(url)
        self.Tables = pd.read_sql_query("show tables", con)
        self.Details = []
        self.Err_table = []

        col0 = self.Tables.columns[0]
        pri_key = {}
        for table in self.Tables[col0]:
            try:
                tb_detail = pd.read_sql_query("DESCRIBE " + table, con)
                tb_detail['tbName'] = table
                self.Details.append(tb_detail)

                key = tb_detail.loc[tb_detail.Key == 'PRI', 'Field']
                if len(key) == 1:
                    key = key[0]
                elif len(key) > 1:
                    key = ','.join(list(key))
                else:
                    key = np.nan
                pri_key[table] = key
            except:
                self.Err_table.append(table)
        key = pd.DataFrame({'priKey': pri_key})
        self.Tables = self.Tables.merge(key, left_on=col0, right_index=True)
        self.Details = pd.concat(self.Details, ignore_index=True)

    def find_tb(self, table_name):
        """找出表名符合正则表达式的所有表\n

        参数:
        ---------
        table_name: str, 可以是普通字符串，也可以是正则表达式描述的表名"""
        col0 = self.Tables.columns[0]
        tb_list = re_search(table_name, self.Tables[col0])
        return tb_list

    def find_col(self, col_name):
        """找出符合正则表达式的所有列名\n

        参数:
        ---------
        col_name : str,可以是普通字符串，也可以是正则表达式描述的列名

        返回:
        ---------
        field_df : dataframe or None, None 表示未找到符合特征的列"""
        col_list = re_search(col_name, self.Details['Field'])
        if col_list:
            logic = self.Details.Field.isin(col_list)
            return self.Details[logic]

    def desc_table(self, table_name):
        """返回指定名称的表的详细信息"""
        logic = self.Details.tbName == table_name
        return self.Details[logic]

    def __repr__(self):
        col = self.Tables.columns[0].split('_')
        db = col[-1]
        tb_n = len(self.Tables)
        col_n = len(self.Details)
        return 'database: {0}, table_num: {1}, col_num: {2}'.format(db, tb_n, col_n)
    

def select_except(table, col, con, alias=None):
    """生成除了 cols 列以外的所有列的 select 代码。当一个表有很多列、除了少数几列外，其他所有列都需要查询时，用此函数 \n
    生成查询的列比手动敲各个列要方便很多。 \n

    参数:
    ----------
    table: 需要 select 的表名\n
    cols: 不希望 select 的字段名\n
    con: MysqlTools 对象"""
    sql_str = "describe " + table
    table_info = con.query(sql_str)
    field = table_info.Field.str.lower()
    if isinstance(col, str):
        col = col.lower()
        logic = field != col  # 剔除单个列
    else:
        col = [i.lower() for i in col]
        logic = ~field.isin(col)  # 剔除多个列组成的序列
    table_filter = table_info[logic]

    if alias is None:  # 表的别名
        alias = table
    field = table_filter.Field.apply(lambda x: alias + "." + x)
    field = ",\n".join(field)
    print("select {0} \nfrom {1}".format(field, table))


def loads_json(json_str):
    """加载 json 字符串为 python 对象。出错就根据格式推断，返回空的 list 或 dict \b

    参数:
    ----------
    json_str: json 格式的字符串"""
    from json import loads
    try:
        x = loads(json_str)
    except:  # 格式错误的话，返回空的 list 或 dict
        x = [] if json_str[:30].strip().startswith('[') else {}
    return x


def dict2casewhen(d_map, col_name):
    """把以字典形式的 {编码：含义} 映射，转换成 sql 的 case when 代码

    参数:
    ------------
    d_map: dict， 以字典形式保存的映射
    col_name: 需要解码的数据库表的字段名

    返回:
    ----------
    无返回值，打印 sql 代码到屏幕"""
    print("case {}".format(col_name))
    for i in d_map:
        print("    when '{0}' then '{1}'".format(i, d_map[i]))
    print("    else 'Unkown' end as {},".format(col_name))


def flatten_multi_index(multi_index, join_str='_'):
    """把 MultiIndex 展平为 1 维。返回一维 Index

    参数:
    -----------
    multi_index: MultiIndex 对象
    join_str: str, 连接第一层和第二层 label 的字符串
    
    返回:
    -----------
    index: 展平后的 index 对象。"""
    label0 = multi_index.get_level_values(0)
    label1 = multi_index.get_level_values(1)
    index = [i + join_str + j for i,j in zip(label0, label1)]
    return pd.Index(index)
	
        
def def_dresult(defm ,df , x ,inplace = True):
    """
    循环数据返回的列表
    df: 数据表
    X: 分组的主键，字符串
    defm: 函数名
    inplace:是否返回值，默认覆盖原数据
    """
    if inplace == True:
        df = df.groupby(x, as_index = False).apply(defm)
    else:
        df.groupby(x, as_index = False).apply(defm)

class GetBirthAge(object):
    """
    处理身份证相关
    解析身份证得到 年龄 性别  生日
    参数：
    --------------------------
    入参:
    身份证ID ，可以是 str  或 list
    apply_date:申请日期 ，如果没有，则默认为 当前日期
    -------------------------------------------------------
    返回：年龄 、性别 、生日
    
    """
    def __init__(self, idcard ,apply_date = None):
        self.id = idcard
        self.apply_date = apply_date
        self.birth_year = list()
        self.birth_month = list()
        self.birth_day = list()
        if isinstance(self.id,(list,pd.Series)):
            for i in self.id:
                if isinstance(i,(str,int)) and len(i) == 18:
                    self.birth_year.append(str(i)[6:10])
                    self.birth_month.append(str(i)[10:12])
                    self.birth_day.append(str(i)[12:14])
                else:
                    self.birth_year.append(np.NaN)
                    self.birth_month.append(np.NaN)
                    self.birth_day.append(np.NaN)                      
        else:
            self.birth_year = self.id[6:10]
            self.birth_month = self.id[10:12]
            self.birth_day = self.id[12:14]
    def get_birthday(self):
        #经过身份证号获取出生日期
        list_num = list()
        if isinstance(self.id,(list,pd.Series)):
            for i in range(len(self.birth_year)):
                if isinstance(self.birth_year[i],str):
                    list_num.append("{0}-{1}-{2}".format(self.birth_year[i], self.birth_month[i], self.birth_day[i]))
                else:
                    list_num.append(None)
            return list_num
        else:            
            birthday = "{0}-{1}-{2}".format(self.birth_year, self.birth_month, self.birth_day)
            return birthday
    def get_sex(self,loc = 17):
        #男生：1 女生：0
        list_num = list()
        if isinstance(self.id,(list,pd.Series)):
            for i in self.id:
                if isinstance(i,(str,int)) and len(i) == 18:
                    if int(i[loc-1:loc]) % 2 == 0:
                        list_num.append('女')
                    else:
                        list_num.append('男')
                else:
                    list_num.append(None)
            return list_num
        else:
            if num % 2 == 0:
                return 0
            else:
                return 1

    def get_age(self):
        #获取年龄
        if isinstance(self.apply_date ,(list ,pd.Series)) and isinstance(self.id,(list,pd.Series)):
            dfi = pd.DataFrame({'idacrd':self.id ,'apply_date': self.apply_date ,'birthday': self.get_birthday()})
            dfi.get_date(['apply_date','birthday'])
            def get_day(row):
                if pd.isnull(row['apply_date']) or pd.isnull(row['birthday']): tmp = np.NaN
                elif int(str(row['apply_date'])[5:7]) <  int(str(row['birthday'])[5:7]):
                    tmp = int(str(row['apply_date'])[:4]) - int(str(row['birthday'])[:4]) - 1
                elif int(str(row['apply_date'])[5:7]) == int(str(row['birthday'])[5:7]) and int(str(row['apply_date'])[8:10]) <= int(str(row['birthday'])[8:10]):
                    tmp = int(str(row['apply_date'])[:4]) - int(str(row['birthday'])[:4]) - 1
                else:
                    tmp = int(str(row['apply_date'])[:4]) - int(str(row['birthday'])[:4])
                return tmp
            dfi['age'] = dfi.apply(get_day ,axis = 1)
            return dfi['age'].to_list()
        elif self.apply_date is None:  
            now = (datetime.now() + timedelta(days=1))
            year = now.year
            month = now.month
            day = now.day
            list_num = list()
            if isinstance(self.id,(list,pd.Series)):
                for i in range(len(self.birth_year)):
                    if isinstance(self.birth_year[i] ,str):
                        if year == int(self.birth_year[i]):
                            list_num.append(0)
                        else:
                            if int(self.birth_month[i]) > month or (int(self.birth_month[i]) == month and int(self.birth_day[i]) > day):
                                list_num.append(year - int(self.birth_year[i]) - 1)
                            else:
                                list_num.append(year - int(self.birth_year[i]))
                    else: list_num.append(np.NaN)
                return list_num
        else:     
            if year == int(self.birth_year):
                return 0
            else:
                if int(self.birth_month) > month or (int(self.birth_month) == month and int(self.birth_day) > day):
                    return year - int(self.birth_year) - 1
                else:
                    return year - int(self.birth_year)
                    
def get_2per(row ,decimal = 2):
    """
    将小数转变为  百分比
    参数：
    ---------------------------
    decimal：需要保留的小数位，默认为2个小数点
    """
    if isinstance(row ,(list,pd.Series)):
        return [str(round(float(i)*100,decimal))+'%' for i in row]
    else:
        return str(round(float(row)*100,decimal))+'%'

def re_change(str_in ,typ = None ,put_out = True):
    """
    字符串的正则替换操作：
    -------------------------------
    str_in: 输入参数
    typ: 输入类型
    ---------------------------------------
    typ == upper: 将字符串中除了单引号之内的所有英文变成大写
    typ == lower: 将字符串中除了单引号之内的所有英文变成小写
    """
    if typ == 'upper':
        clo = re.findall(r"'([^']*)'",str_in)
        str_in = str_in.upper()
        col = re.findall(r"'([^']*)'",str_in)
        for i in range(len(clo)):
            str_in = str_in.replace(col[i] ,clo[i])
    elif typ == 'lower':
        clo = re.findall(r"'([^']*)'",str_in)
        str_in = str_in.lower()
        col = re.findall(r"'([^']*)'",str_in)
        for i in range(len(clo)):
            str_in = str_in.replace(col[i] ,clo[i])
    else:
        pass
    if put_out:
        print(str_in)
    else:
        return str_in


def date_diff(a ,b ,dtype = 'month'):
    """
    计算两个日期的差
    ---------------------------
    dtype:类型，默认为月份差， 可以是 year ，day
    默认为：a - b的结果
    -------------------------
    返回结果为 差值
    """
    def get_date(x):
        year = int(str(x).split('-')[0])
        month = int(str(x).split('-')[1])
        day = int(str(x).split('-')[2])
        Sum = 0
        total_month = (31 ,28 ,31 ,30 ,31 ,30 ,31 ,31 ,30 ,31 ,30 ,31)
        Sum = sum(total_month[:month - 1]) + day
        percent = (Sum/365) * 100
        if year % 400 == 0 or (year%4 == 0 and year % 100 !=0):
            if month > 2:
                Sum+=1
            percent = (Sum/366) * 100
        percent = round(percent/100 ,4)
        return percent
    
    if a > b:
        km = a.year - b.year
        if km == 0:
            year = get_date(a) - get_date(b)
        else :
            year = km - get_date(b) + get_date(a)
    else:
        km = b.year - a.year
        if km == 0:
            year = get_date(b) - get_date(a)
        else :
            year = km - get_date(a) + get_date(b)   
            
    month = int(year*12)
    if dtype == 'year':
        return round(year,4)
    elif dtype == 'month':
        return month
    elif dtype == 'day':
        return (a-b).days
        