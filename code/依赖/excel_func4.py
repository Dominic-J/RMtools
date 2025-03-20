# -*- coding: utf-8 -*-
# @Author  : zhangjiatao
# @Time    : 2024/2/4 12:17
# @Function: excel格式化函数
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import re

def find_target_positions(sheet,target_value):
    positions = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == target_value:
                position = f"{get_column_letter(cell.column)}{cell.row}"
                positions.append(position)

    return positions


def find_empty_cell(sheet, coordinate):
    # 获取列号和起始行号
    # column = coordinate[0]
    column = re.sub(u"([^\u0041-\u007a])","",coordinate)
    # start_row = int(coordinate[1:])
    start_row = int(re.sub(u"([^\u0030-\u0039])","",coordinate))


    # 获取指定列的单元格范围
    cells_in_column = sheet[column]


    for cell in cells_in_column[start_row - 1:]:
        if cell.value is None:
            return f"{column}{cell.row}"


    # 如果没有找到空值单元格，则返回None
    return f"{column}{cells_in_column[start_row - 1:-1][-1].row+2}"



def set_color_scale(sheet,start_coordinate, end_coordinate,n):
    # 获取起始和结束列号、行号
    start_column = re.sub(u"([^\u0041-\u007a])","",start_coordinate)
    # start_row = int(re.sub(u"([^\u0030-\u0039])","",start_coordinate))
    # lift下的第一行是missing，不进行颜色区分
    start_row = int(re.sub(u"([^\u0030-\u0039])","",start_coordinate))+n+1
    end_column = re.sub(u"([^\u0041-\u007a])","",end_coordinate)
    end_row = int(re.sub(u"([^\u0030-\u0039])","",end_coordinate))

    # 获取起始和结束列的字母表示
    # start_col_letter = get_column_letter(ord(start_column.upper()) - 64)
    # end_col_letter = get_column_letter(ord(end_column.upper()) - 64)
    start_col_letter = start_column
    end_col_letter = end_column

    # 创建色阶规则对象
    color_scale_rule = ColorScaleRule(
        start_type="min",
        start_color=Color(rgb="63BE7B"),
        mid_type='percentile',
        mid_value=50,
        mid_color='FFEB84',
        end_type="max",
        end_color=Color(rgb="F8696B")
    )

    # 应用色阶规则
    if start_row < end_row:
        range_string = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        sheet.conditional_formatting.add(range_string, color_scale_rule)


def excel_value_format(sheet,column,row,format_type):
    from openpyxl.styles import numbers
    # 获取目标单元格对象
    target_cell = sheet[column + str(row)]

    # 根据格式类型设置单元格格式
    if format_type == 'percentage':
        # 设置为百分比格式
        target_cell.number_format = numbers.FORMAT_PERCENTAGE_00
    elif format_type == 'decimal':
        # 设置为保留两位小数的数值格式
        target_cell.number_format = numbers.FORMAT_NUMBER_00

def highlight_max_values(sheet, start_coordinate, end_coordinate):
    from openpyxl.styles import PatternFill
    # 获取起始和结束列号、行号
    start_col = re.sub(u"([^\u0041-\u007a])","",start_coordinate)
    start_row = int(re.sub(u"([^\u0030-\u0039])","",start_coordinate))
    end_col = re.sub(u"([^\u0041-\u007a])","",end_coordinate)
    end_row = int(re.sub(u"([^\u0030-\u0039])","",end_coordinate))

    # 创建填充格式
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 找到指定范围内的最大值
    max_value = None
    for row in range(start_row+1, end_row-1):
        cell = sheet[f"{start_col}{row}"]
        if max_value is None or cell.value > max_value:
            max_value = cell.value

    # 标记最大值单元格的颜色
    for row in range(start_row+1, end_row-1):
        cell = sheet[f"{start_col}{row}"]
        if cell.value == max_value:
            cell.fill = fill


if __name__ == "__main__":
    filename='D:/shortcut/临时文件/926临时文件/excel_测试/test.xlsx'
    keyword='P3'
    res = find_empty_cell(filename,keyword)
    print(res)